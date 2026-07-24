"""
download_censo_2022.py — EL (Extract & Load) do Censo Demográfico 2022

Downloads CSVs from IBGE FTP, merges all themes into a single GeoParquet
with census tract geometries, ready for ``ambx.demographics``.

Usage:
    python scripts/extract/download_censo_2022.py

Output:
    data/raw/censo_2022/
    ├── censo_2022.gpkg          ← single GeoParquet (all variables + geometry)
    └── metadados.json           ← variable dictionary
"""

from __future__ import annotations

import json
import zipfile
from pathlib import Path

import geopandas as gpd
import openpyxl
import pandas as pd
import requests

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

URL_BASE = "https://ftp.ibge.gov.br/Censos/Censo_Demografico_2022"
PROJ_ROOT = Path(__file__).resolve().parents[2]
DATA_RAW = PROJ_ROOT / "data" / "raw" / "censo_2022"
CACHE = Path(__file__).resolve().parent / "cache_ibge"

TRACT_ID = "CD_SETOR"

# ---------------------------------------------------------------------------
# File URLs on IBGE FTP
# ---------------------------------------------------------------------------

FILES = {
    "basico": (
        f"{URL_BASE}/Agregados_por_Setores_Censitarios/Agregados_por_Setor_csv/"
        "Agregados_por_setores_basico_BR_20260520.zip"
    ),
    "literacy": (
        f"{URL_BASE}/Agregados_por_Setores_Censitarios/Agregados_por_Setor_csv/"
        "Agregados_por_setores_alfabetizacao_BR.zip"
    ),
    "race": (
        f"{URL_BASE}/Agregados_por_Setores_Censitarios/Agregados_por_Setor_csv/"
        "Agregados_por_setores_cor_ou_raca_BR.zip"
    ),
    "demography": (
        f"{URL_BASE}/Agregados_por_Setores_Censitarios/Agregados_por_Setor_csv/"
        "Agregados_por_setores_demografia_BR.zip"
    ),
    "income": (
        f"{URL_BASE}/Agregados_por_Setores_Censitarios_Rendimento_do_Responsavel/"
        "Agregados_por_setores_renda_responsavel_BR_20260508_csv.zip"
    ),
}

DICTIONARIES = {
    "census": (
        f"{URL_BASE}/Agregados_por_Setores_Censitarios/"
        "dicionario_de_dados_agregados_por_setores_censitarios_20260520.xlsx"
    ),
    "income": (
        f"{URL_BASE}/Agregados_por_Setores_Censitarios_Rendimento_do_Responsavel/"
        "dicionario_de_dados_renda_responsavel_20260508.xlsx"
    ),
}

# ---------------------------------------------------------------------------
# Variables of interest per theme
# ---------------------------------------------------------------------------

COLUMNS: dict[str, list[str]] = {
    # Basico usa v minúsculo, os demais V maiúsculo
    "basico": ["v0001", "v0002"],
    "literacy": [
        "V00644", "V00645", "V00646", "V00647", "V00648",
        "V00649", "V00650", "V00651", "V00652", "V00653",
        "V00654", "V00655", "V00656",
        "V00748", "V00749", "V00750", "V00751", "V00752",
        "V00753", "V00754", "V00755", "V00756", "V00757",
        "V00758", "V00759", "V00760",
        "V00900", "V00901", "V00984", "V00985",
    ],
    "race": [
        "V01317", "V01318", "V01319", "V01320", "V01321",
        "V01322", "V01323", "V01324", "V01325", "V01326",
        "V01327", "V01328", "V01329", "V01330", "V01331",
        "V01332", "V01333", "V01334", "V01335", "V01336",
        "V01372", "V01373", "V01374", "V01375", "V01376",
        "V01377", "V01378", "V01379", "V01380", "V01381",
        "V01382", "V01383", "V01384", "V01385", "V01386",
        "V01387", "V01388", "V01389", "V01390", "V01391",
    ],
    "demography": [
        "V01006", "V01007", "V01008",
        "V01009", "V01010", "V01011", "V01012", "V01013",
        "V01014", "V01015", "V01016", "V01017", "V01018",
        "V01019",
        "V01020", "V01021", "V01022", "V01023", "V01024",
        "V01025", "V01026", "V01027", "V01028", "V01029",
        "V01030",
        "V01031", "V01032", "V01033", "V01034", "V01035",
        "V01036", "V01037", "V01038", "V01039", "V01040",
        "V01041",
    ],
    "income": ["V06001", "V06002", "V06004", "V06006"],
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _download(url: str, dest: Path) -> None:
    if dest.exists():
        print(f"  \u2714 {dest.name}")
        return
    print(f"  \u2b07 {dest.name} ...", end=" ", flush=True)
    r = requests.get(url, stream=True, timeout=300)
    r.raise_for_status()
    with open(dest, "wb") as f:
        for chunk in r.iter_content(chunk_size=8192):
            f.write(chunk)
    print(f"{dest.stat().st_size / 1e6:.1f} MB")


def _normalize_tract_id(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        if col.upper() == TRACT_ID:
            if col != TRACT_ID:
                df = df.rename(columns={col: TRACT_ID})
            break
    return df


def _read_zip(zip_path: Path) -> pd.DataFrame:
    with zipfile.ZipFile(zip_path) as z:
        csv = next(n for n in z.namelist() if n.endswith(".csv"))
        return pd.read_csv(z.open(csv), sep=";", encoding="latin1", low_memory=False)


def _load_metadata(xlsx: Path, sheet: str) -> dict:
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[sheet]
    meta = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = [str(c).strip() if c else "" for c in row]
        var_idx = 2 if len(vals) >= 4 else 1
        desc_idx = 3 if len(vals) >= 4 else 2
        if len(vals) > max(var_idx, desc_idx) and vals[var_idx].startswith("V"):
            meta[vals[var_idx]] = vals[desc_idx]
    return meta


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    print("=" * 60)
    print("  EL — Censo Demografico 2022")
    print("  Download, merge and save single GeoParquet with geometry")
    print("=" * 60)

    DATA_RAW.mkdir(parents=True, exist_ok=True)
    CACHE.mkdir(exist_ok=True)

    # 1. Dictionaries
    print("\n[Dictionaries]")
    for name, url in DICTIONARIES.items():
        _download(url, CACHE / url.split("/")[-1])

    dic_census = CACHE / DICTIONARIES["census"].split("/")[-1]
    dic_income = CACHE / DICTIONARIES["income"].split("/")[-1]

    # 2. Data files
    print("\n[Data files]")
    for name, url in FILES.items():
        _download(url, CACHE / url.split("/")[-1])

    # 3. Census tract geometries (malha com atributos do IBGE)
    print("\n[Tract geometries]")
    geo_url = (
        "https://ftp.ibge.gov.br/Censos/Censo_Demografico_2022/"
        "Agregados_por_Setores_Censitarios/"
        "malha_com_atributos/setores/gpkg/BR/BR_setores_CD2022.gpkg"
    )
    geo_dest = CACHE / "BR_setores_CD2022.gpkg"
    _download(geo_url, geo_dest)

    print("  Loading...", end=" ", flush=True)
    tracts = gpd.read_file(geo_dest)
    for col in tracts.columns:
        if col.upper() == TRACT_ID:
            if col != TRACT_ID:
                tracts = tracts.rename(columns={col: TRACT_ID})
            break
    tracts[TRACT_ID] = tracts[TRACT_ID].astype(str)
    # Descartar colunas da malha que conflitam com os dados (v0001-v0009)
    cols_to_drop = [c for c in tracts.columns if c.lower().startswith("v000")]
    if cols_to_drop:
        tracts = tracts.drop(columns=cols_to_drop)
    print(f"{len(tracts):,} tracts")

    # 4. Merge all themes
    print("\n[Merge]")
    merged = None

    for theme in ["basico", "literacy", "race", "demography", "income"]:
        url = FILES[theme]
        path = CACHE / url.split("/")[-1]
        cols = COLUMNS[theme]

        print(f"  {theme}...", end=" ", flush=True)
        df = _read_zip(path)
        df = _normalize_tract_id(df)
        df[TRACT_ID] = df[TRACT_ID].astype(str)

        existing = [c for c in cols if c in df.columns]
        df = df[[TRACT_ID] + existing].set_index(TRACT_ID)

        if merged is None:
            merged = df
        else:
            merged = merged.join(df, how="outer")
        print(f"{len(existing)} columns")

    # 5. Join geometry
    print("\n[Joining geometry]...", end=" ", flush=True)
    gdf = tracts.merge(merged.reset_index(), on=TRACT_ID, how="inner")
    print(f"{len(gdf):,} tracts with geometry")

    # ── 5b. Corrigir dtypes das colunas V... ──────────────────────
    # O IBGE codifica valores censurados como "X" e a leitura com
    # low_memory=False os trata como string. Convertemos para float,
    # tratando "X" como NaN.
    print("\n[Type coercion]...", end=" ", flush=True)
    v_cols = [c for c in gdf.columns if c.startswith("V")]
    for c in v_cols:
        gdf[c] = pd.to_numeric(gdf[c], errors="coerce")
    print(f"{len(v_cols)} columns converted to float64")

    # 6. Save GeoParquet
    out = DATA_RAW / "censo_2022.geoparquet"
    print(f"\n[Saving] {out} ...", end=" ", flush=True)
    gdf.to_parquet(out, index=False)
    size = out.stat().st_size / 1e6
    print(f"{size:.1f} MB")

    # 7. Metadata
    print("[Metadata]...", end=" ", flush=True)
    meta: dict[str, object] = {}
    meta["census"] = _load_metadata(dic_census, "Dicionário não PCT")
    meta["basico"] = _load_metadata(dic_census, "Dicionário Básico")
    meta["income"] = _load_metadata(dic_income, "Dicionário Renda Responsável")
    with open(DATA_RAW / "metadados.json", "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    print("ok")

    # 8. Summary
    print(f"\n{'=' * 60}")
    print("  Done!")
    print(f"  Folder: {DATA_RAW}")
    print(f"    censo_2022.gpkg  ({size:.1f} MB, {len(gdf.columns)} columns)")
    print(f"    metadados.json")
    print(f"  Cache: {CACHE}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
